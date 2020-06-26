import os
import sys
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def e2p(datapath):

    data = pd.ExcelFile(datapath)
    
    return data

def get_sheet_names(data): 

    wb = load_workbook(data)
    sheetnames = wb.sheetnames
    
    return sheetnames


def get_mouseID(datapath):

    mouse = datapath.split('_')[1]
    celltemp = datapath.split('_')[3]
    cellraw = r'{}'.format(celltemp)
    cell = cellraw.split("\\")[0]
    mouseID = mouse + ': ' + cell

    return mouseID

#def get_recording_date(datapath):

#    recording_date = datapath.split('_A')[0].split("s\\")[1]
    
#    return recording_date

def Total_dendrite_length(data,sheetnames): 

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_length_total = df7.iloc[2,13]
        
        return dendrite_length_total
    else: 
        return np.nan
    

def Av_dendrite_length(data,sheetnames): 

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_length_average = df7.iloc[2,14]
        
        return dendrite_length_average
    else: 
        return np.nan

def Total_dendrite_SA(data,sheetnames):

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_SA = df7.iloc[2,17]
        
        return dendrite_SA
    else: 
        return np.nan


def Dendrite_av_polar_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 

        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_av_polar_angle = df9.iloc[:,2].mean()

        return dendrite_av_polar_angle 
        
    else: 
        return np.nan

def Dendrite_av_local_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 
        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_av_local_angle = df9.iloc[:,3].mean()

        return dendrite_av_local_angle   
    else:
        return np.nan 

def Dendrite_sholl_number(data,sheetnames):

    if 'Sholl - Dendrites' in sheetnames: 

        df12 = pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_sholl_number = math.ceil(len(df12.iloc[:])/5)

        return dendrite_sholl_number            
    else: 
        return np.nan
    
def Dendrite_sholl_length_density(data,sheetnames):

    sholl_number = Dendrite_sholl_number(data,sheetnames)
    dendrite_length_total = Total_dendrite_length(data,sheetnames)
    dendrite_sholl_length_density= dendrite_length_total/sholl_number 

    return dendrite_sholl_length_density

        
def make_param_df(datapath):

    data = e2p(datapath)
    sheetnames = get_sheet_names(data)

    Parameter_dict ={
                    'Total_dendrite_length': Total_dendrite_length(data,sheetnames=sheetnames), 
                    'Average_dendrite_length': Av_dendrite_length(data,sheetnames=sheetnames), 
                    'Dendrite_SA': Total_dendrite_SA(data,sheetnames=sheetnames), 
                    'Dendrite_polar_angle': Dendrite_av_polar_angle(data,sheetnames=sheetnames),  
                    'Dendrite_local_angle': Dendrite_av_local_angle(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_length_density': Dendrite_sholl_length_density(data,sheetnames=sheetnames),
}

    df_params = pd.DataFrame.from_dict(Parameter_dict, orient = 'index')
    df_params.columns = [get_mouseID(datapath)]
    return (df_params)