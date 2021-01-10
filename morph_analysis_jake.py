import os
import sys
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def e2p(datapath):

    data = pd.ExcelFile(datapath)
    
    return data

def get_sheet_names(data): 

    wb = load_workbook(data)
    sheetnames = wb.sheetnames
    
    return sheetnames


def get_mouseID(datapath):

    mouse = datapath.split('_')[1]
    celltemp = datapath.split('_')[3]
    cellraw = r'{}'.format(celltemp)
    cell = cellraw.split("\\")[0]
    cell=cell.replace("0","")
    mouseID = mouse + ': ' + cell
    

    return mouseID

#def get_recording_date(datapath):

#    recording_date = datapath.split('_A')[0].split("s\\")[1]
    
#    return recording_date
       
def Axon_node_total(data,sheetnames): 
    

    if 'Neuron Summary' in sheetnames:

       df1 = pd.read_excel(data, 'Neuron Summary')
       axon_node_total = df1.iloc[1,2]
       return axon_node_total
    else: 
        return np.nan 

def Total_axon_length(data,sheetnames):
    
    
    if 'Neuron Summary' in sheetnames: 

        df1 = pd.read_excel(data, 'Neuron Summary')
        axon_length_total = df1.iloc[1,13]
        
        return axon_length_total
    else: 
        return np.nan 

def Total_axon_SA(data,sheetnames):
    
    
    if 'Neuron Summary' in sheetnames:
        df1 = pd.read_excel(data, 'Neuron Summary')
        axon_SA = df1.iloc[1,17]
        return axon_SA
    else: 
        return np.nan     

def Axon_len_SA_ratio(data,sheetnames):
    
    
    if 'Neuron Summary' in sheetnames:
        
        axon_length_total = Total_axon_length(data, sheetnames)
        axon_SA = Total_axon_SA(data,sheetnames)
        axon_len_SA_ratio = axon_length_total/axon_SA
        
        return axon_len_SA_ratio
    else: 
        return np.nan 

def Highest_order_axon_segment(data,sheetnames):
    

    if 'Each Tree - Axons' in sheetnames:
        df2 = pd.read_excel(data, 'Each Tree - Axons')
        highest_order_axon_segment = df2.iloc[:,2].max()
        return highest_order_axon_segment
    else:
        return np.nan 

def Axon_av_polar_angle(data,sheetnames):

    # This is extracting the mean Planar Angle - pls correct if incorrect value 
    

    if 'Branch Angle-Axon' in sheetnames: 
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_av_polar_angle = df3.iloc[:,2].mean()
        return axon_av_polar_angle 
    else: 
        return np.nan

def Axon_std_polar_angle(data,sheetnames):

    # as above 
    
    
    if 'Branch Angle-Axon' in sheetnames: 
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_std_polar_angle = df3.iloc[:,2].std()
        return axon_std_polar_angle 
    else: 
        return np.nan 

def Axon_av_local_angle(data,sheetnames):
    

    if 'Branch Angle-Axon' in sheetnames:
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_av_local_angle = df3.iloc[:,3].mean()
        return axon_av_local_angle 
    else: 
        return np.nan
    
    
def Axon_std_local_angle(data,sheetnames):
    
    
    if 'Branch Angle-Axon' in sheetnames:
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_std_local_angle = df3.iloc[:,3].std()
        return axon_std_local_angle 
    else: 
        return np.nan

def Axon_av_spline_angle(data,sheetnames):
    

    if 'Branch Angle-Axon' in sheetnames:
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_av_spline_angle = df3.iloc[:,4].mean()
        return axon_av_spline_angle 
    else: 
        return np.nan

def Axon_std_spline_angle(data,sheetnames):
    
    
    if 'Branch Angle-Axon' in sheetnames:
        df3 = pd.read_excel(data, 'Branch Angle-Axon')
        axon_std_spline_angle = df3.iloc[:,4].std()
        
        return axon_std_spline_angle 
    else: 
        return np.nan

def Axon_segment_av_tortuosity(data,sheetnames):
    
    
    if 'Segment - Axons' in sheetnames:
        df4 = pd.read_excel(data, 'Segment - Axons')
        axon_segment_av_tortuosity = df4.iloc[:,3].mean()

        return axon_segment_av_tortuosity
    else: 
        return np.nan

def Axon_segment_std_tortuosity(data,sheetnames):  
    

    if 'Segment - Axons' in sheetnames:
        df4 = pd.read_excel(data, 'Segment - Axons')
        axon_segment_std_tortuosity = df4.iloc[:,3].std()

        return axon_segment_std_tortuosity
    else: 
        return np.nan

def Axon_segment_av_length(data,sheetnames): 
    
    
    if 'Segment - Axons' in sheetnames:
        df4 = pd.read_excel(data, 'Segment - Axons')
        axon_length_total = Total_axon_length(data,sheetnames)
        axon_segment_av_length = axon_length_total/len(df4.iloc[:])

        return axon_segment_av_length 
    else: 
        return np.nan

def Axon_segment_std_length(data,sheetnames): 
    
    
    if 'Segment - Axons' in sheetnames:
        df4 = pd.read_excel(data, 'Segment - Axons')
        axon_segment_std_length = df4.iloc[:,2].std()

        return axon_segment_std_length 
    else: 
        return np.nan

def Axon_nodes_av_tortuosity(data,sheetnames):
    

    if 'Nodes - Axons' in sheetnames:
        df5 = pd.read_excel(data, 'Nodes - Axons')
        axon_segment_av_tortuosity = df5.iloc[:,4].mean()

        return axon_segment_av_tortuosity
    else: 
        return np.nan

def Axon_nodes_std_tortuosity(data,sheetnames):
    

    if 'Nodes - Axons' in sheetnames:
        df5 = pd.read_excel(data, 'Nodes - Axons')
        axon_nodes_std_tortuosity = df5.iloc[:,4].std()

        return axon_nodes_std_tortuosity
    else: 
        return np.nan

def Axon_sholl_number(data,sheetnames):
    
    # this counts the number of shells w radius increasing in steps of 100 microns 

    if 'Sholl - Axons' in sheetnames:
        df6 = pd.read_excel(data, 'Sholl - Axons')
        axon_sholl_number = math.ceil(len(df6.iloc[:])/10)

        return axon_sholl_number 
    else: 
        return np.nan

def Axon_sholl_length_100(data,sheetnames):
    
    
    if 'Sholl - Axons' in sheetnames:
        df6 = pd.read_excel(data, 'Sholl - Axons')
        axon_length_total = Total_axon_length(data, sheetnames)
        axon_sholl_length_100 = sum(df6.iloc[0:10,2])/axon_length_total

        return axon_sholl_length_100 
    else: 
        return np.nan

def Axon_sholl_length_200(data,sheetnames):
    
    
    if 'Sholl - Axons' in sheetnames:
        df6 = pd.read_excel(data, 'Sholl - Axons')
        axon_length_total = Total_axon_length(data, sheetnames)
        axon_sholl_length_200 = sum(df6.iloc[11:20,2])/axon_length_total

        return axon_sholl_length_200
    else: 
        return np.nan

def Axon_sholl_length_300(data,sheetnames):
    
    if 'Sholl - Axons' in sheetnames:
        df6 = pd.read_excel(data, 'Sholl - Axons')
        axon_length_total = Total_axon_length(data, sheetnames)
        axon_sholl_length_300 = sum(df6.iloc[21:30,2])/axon_length_total

        return axon_sholl_length_300
    else: 
        return np.nan

def Axon_sholl_length_density(data,sheetnames):
    
    
    if 'Sholl - Axons' in sheetnames:
        sholl_number = Axon_sholl_number(data, sheetnames)
        axon_length_total = Total_axon_length(data, sheetnames)
        axon_sholl_length_density= axon_length_total/sholl_number 

        return axon_sholl_length_density
    else: 
        return np.nan

def Axon_sholl_node_density(data,sheetnames):
    
    
    if 'Sholl - Axons' in sheetnames:
        sholl_number = Axon_sholl_number(data, sheetnames)
        axon_node_total = Axon_node_total(data, sheetnames)
        axon_sholl_node_density = axon_node_total/sholl_number 

        return axon_sholl_node_density
    else: 
        return np.nan

def Dendrite_number(data,sheetnames):

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_number = df7.iloc[2,1]
        
        return dendrite_number
    else: 
        return np.nan

def Dendrite_node_total(data,sheetnames): 

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_node_total = df7.iloc[2,2]
        
        return dendrite_node_total
    else: 
        return np.nan
    

def Total_dendrite_length(data,sheetnames): 

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_length_total = df7.iloc[2,13]
        
        return dendrite_length_total
    else: 
        return np.nan
    

def Av_dendrite_length(data,sheetnames): 

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_length_average = df7.iloc[2,14]
        
        return dendrite_length_average
    else: 
        return np.nan

def Total_dendrite_SA(data,sheetnames):

    if 'Neuron Summary' in sheetnames: 

        df7 = pd.read_excel(data, 'Neuron Summary')
        dendrite_SA = df7.iloc[2,17]
        
        return dendrite_SA
    else: 
        return np.nan

def Dendrite_len_SA_ratio(data,sheetnames):
    
    
    dendrite_length_total = Total_dendrite_length(data,sheetnames)
    dendrite_SA = Total_dendrite_SA(data,sheetnames)
    dendrite_len_SA_ratio = dendrite_length_total/dendrite_SA
    
    return dendrite_len_SA_ratio

def Highest_order_dendrite_segment(data,sheetnames):

    if 'Each Tree - Dendrites' in sheetnames: 

        df8 = pd.read_excel(data, 'Each Tree - Dendrites')
        highest_order_dendrite_segment = df8.iloc[:,2].max()
        
        return highest_order_dendrite_segment
    else: 
        return np.nan

def Dendrite_av_polar_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 

        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_av_polar_angle = df9.iloc[:,2].mean()

        return dendrite_av_polar_angle 
        
    else: 
        return np.nan

  

def Dendrite_std_polar_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 

        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_std_polar_angle = df9.iloc[:,2].std()

        return dendrite_std_polar_angle 
        
    else: 
        return np.nan


def Dendrite_av_local_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 
        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_av_local_angle = df9.iloc[:,3].mean()

        return dendrite_av_local_angle   
    else:
        return np.nan 

def Dendrite_std_local_angle(data,sheetnames):
    if 'Branch Angle-Dendrite' in sheetnames: 
        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_std_local_angle = df9.iloc[:,3].std()

        return dendrite_std_local_angle  
    else:
        return np.nan  


def Dendrite_av_spline_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 

        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_av_spline_angle = df9.iloc[:,4].mean()

        return dendrite_av_spline_angle  
    else: 
        return np.nan  

def Dendrite_std_spline_angle(data,sheetnames):

    if 'Branch Angle-Dendrite' in sheetnames: 

        df9 = pd.read_excel(data, 'Branch Angle-Dendrite')
        dendrite_std_spline_angle = df9.iloc[:,4].std()

        return dendrite_std_spline_angle 
    else: 
        return np.nan  


def Dendrite_segment_av_tortuosity(data,sheetnames):

    if 'Segment - Dendrites' in sheetnames: 

        df10 = pd.read_excel(data, 'Segment - Dendrites')
        dendrite_segment_av_tortuosity = df10.iloc[:,3].mean()

        return dendrite_segment_av_tortuosity      
    else: 
        return np.nan  
    

def Dendrite_segment_std_tortuosity(data,sheetnames):

    if 'Segment - Dendrites' in sheetnames: 

        df10 = pd.read_excel(data, 'Segment - Dendrites')
        dendrite_segment_std_tortuosity = df10.iloc[:,3].std()

        return dendrite_segment_std_tortuosity     
    else: 
        return np.nan 


def Dendrite_segment_av_length(data,sheetnames): 

    if 'Segment - Dendrites' in sheetnames: 

        df10 = pd.read_excel(data, 'Segment - Dendrites')
        dendrite_length_total = Total_dendrite_length(data,sheetnames)
        dendrite_segment_av_length = dendrite_length_total/len(df10.iloc[:])

        return dendrite_segment_av_length
            
    else: 
        return np.nan 


def Dendrite_segment_std_length(data,sheetnames): 

    if 'Segment - Dendrites' in sheetnames: 

        df10 = pd.read_excel(data, 'Segment - Dendrites')
        dendrite_segment_std_length = df10.iloc[:,2].std()

        return dendrite_segment_std_length         
    else: 
        return np.nan 

def Dendrite_nodes_av_tortuosity(data,sheetnames):

    if 'Nodes - Dendrites' in sheetnames: 

        df11 = pd.read_excel(data, 'Nodes - Dendrites')
        dendrite_segment_av_tortuosity = df11.iloc[:,4].mean()

        return dendrite_segment_av_tortuosity       
    else: 
        return np.nan 

def Dendrite_nodes_std_tortuosity(data,sheetnames):

    if 'Nodes - Dendrites' in sheetnames: 

        df11 = pd.read_excel(data, 'Nodes - Dendrites')
        dendrite_nodes_std_tortuosity = df11.iloc[:,4].std()

        return dendrite_nodes_std_tortuosity          
    else: 
        return np.nan 

def Dendrite_sholl_number(data,sheetnames):

    if 'Sholl - Dendrites' in sheetnames: 

        df12 = pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_sholl_number = math.ceil(len(df12.iloc[:])/5)

        return dendrite_sholl_number            
    else: 
        return np.nan

def Dendrite_sholl_length_50(data,sheetnames):

    if 'Sholl - Dendrites' in sheetnames: 

        df12 = pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_length_total = Total_dendrite_length(data,sheetnames)
        dendrite_sholl_length_50 = sum(df12.iloc[0:5,2])/dendrite_length_total

        return dendrite_sholl_length_50           
    else: 
        return np.nan


def Dendrite_sholl_length_100(data,sheetnames):

    if 'Sholl - Dendrites' in sheetnames: 

        df12 = pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_length_total = Total_dendrite_length(data,sheetnames)
        dendrite_sholl_length_100 = sum(df12.iloc[6:10,2])/dendrite_length_total
        return dendrite_sholl_length_100           
    else: 
        return np.nan

def Dendrite_sholl_length_150(data,sheetnames):

    if 'Sholl - Dendrites' in sheetnames: 
        
        df12 = pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_length_total = Total_dendrite_length(data,sheetnames)
        dendrite_sholl_length_150 = sum(df12.iloc[11:15,2])/dendrite_length_total
        
        return dendrite_sholl_length_150           
    else: 
        return np.nan


def Dendrite_sholl_length_density(data,sheetnames):

    sholl_number = Dendrite_sholl_number(data,sheetnames)
    dendrite_length_total = Total_dendrite_length(data,sheetnames)
    dendrite_sholl_length_density= dendrite_length_total/sholl_number 

    return dendrite_sholl_length_density

def Dendrite_sholl_node_density(data,sheetnames):

    sholl_number = Dendrite_sholl_number(data,sheetnames)
    dendrite_node_total = Dendrite_node_total(data,sheetnames)
    dendrite_sholl_node_density = dendrite_node_total/sholl_number 

    return dendrite_sholl_node_density

def Dendrite_sholl_number_10(data,sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
        
        df12=pd.read_excel(data,'Sholl - Dendrites')
        dendrite_intersections_10 = df12.iloc[1,1]
        
        return dendrite_intersections_10
    else:
        return np.nan
    
def Dendrite_sholl_number_50(data,sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
        
        df12=pd.read_excel(data,'Sholl - Dendrites')
        try:
            dendrite_intersections_50 = df12.iloc[5,1]
        
        except IndexError:
            dendrite_intersections_50=0
        return dendrite_intersections_50
    else:
        return np.nan

def Dendrite_sholl_number_100(data,sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
        
        df12=pd.read_excel(data,'Sholl - Dendrites')
        try:
            dendrite_intersections_100 = df12.iloc[10,1]
        except IndexError:
            dendrite_intersections_100=0
        
        return dendrite_intersections_100
    else:
        return np.nan

def Dendrite_sholl_number_200(data,sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
        
        df12=pd.read_excel(data,'Sholl - Dendrites')
        try:
            dendrite_intersections_200 = df12.iloc[20,1]
        except IndexError:
            dendrite_intersections_200=0
        return dendrite_intersections_200
    else:
        return np.nan

def Dendrite_sholl_number_300(data,sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
        
        df12=pd.read_excel(data,'Sholl - Dendrites')
        try:
            dendrite_intersections_300 = df12.iloc[30,1]
        except IndexError:
            dendrite_intersections_300=0
        return dendrite_intersections_300
    else:
        return np.nan
    
# The following functions are for extracing soma-related variables 
# Soma data is split into multiple slices and each slice analysed separately by Neurolucida --> may be worth checking if we can change this 

def Soma_perimeter(data,sheetnames):

    # written to extract maximum perimeter from list of slices - pls alter if incorrect 

    if 'Cell Bodies' in sheetnames: 
        df13 = pd.read_excel(data, 'Cell Bodies')
        soma_perimeter = df13.iloc[:,0].max()
        return soma_perimeter              
    else: 
        
        return np.nan
    

def Soma_SA(data,sheetnames):

    # Cell Bodie not a typo - sheet named this automatically by Neurolucida Ex. 

    if '3D Contour Summary - Cell Bodie' in sheetnames: 
        df14 = pd.read_excel(data, '3D Contour Summary - Cell Bodie')
        soma_SA = df14.iloc[0,3]
        return soma_SA           
    else:  
        return np.nan
    
def Soma_volume(data,sheetnames):

    # As above

    if '3D Contour Summary - Cell Bodies' in sheetnames: 

        df14 = pd.read_excel(data, '3D Contour Summary - Cell Bodie')
        soma_volume = df14.iloc[0,2]

        return soma_volume              
    else: 
        
        return np.nan


def Soma_aspect_ratio(data,sheetnames):

    if 'Cell Bodies' in sheetnames: 

        df13 = pd.read_excel(data, 'Cell Bodies')
        soma_aspect_ratio = df13.iloc[:,2].max()/df13.iloc[:,3].min()

        return soma_aspect_ratio              
    else: 
        
        return np.nan


def Soma_compactness_factor(data,sheetnames):

    if 'Cell Bodies' in sheetnames: 

        df13 = pd.read_excel(data, 'Cell Bodies')
        area = Soma_SA(data,sheetnames)
        max_diameter = df13.iloc[:,2].max()
        soma_compactness_factor = (((4/math.pi)*area)**0.5)/max_diameter

        return soma_compactness_factor              
    else: 
        
        return np.nan

def Somatic_form_factor(data,sheetnames):

    perimeter = Soma_perimeter(data,sheetnames)
    area = Soma_SA(data,sheetnames)
    somatic_form_factor = (4*math.pi*area)/(perimeter**2)

    return somatic_form_factor 

def Somatic_roundness(data,sheetnames):

    if 'Cell Bodies' in sheetnames: 

        area = Soma_SA(data,sheetnames)
        df13 = pd.read_excel(data, 'Cell Bodies')
        max_diameter = df13.iloc[:,2].max()
        somatic_roundness = (4*area)/(math.pi*max_diameter**2)

        return somatic_roundness
                
    else: 
        
        return np.nan

def Dendritic_polarity(data,sheetnames):

    if 'Polar Histogram - Dendrites' in sheetnames:
            
        df10 = pd.read_excel(data, 'Polar Histogram - Dendrites')
        dendritic_polarity = df10.iloc[:,2].std()/df10.iloc[:,2].mean()
        
        return dendritic_polarity
    
    else:
            
            return np.nan

def dendrite_max_extent(data, sheetnames):
    
    if 'Sholl - Dendrites' in sheetnames:
    
        df10=pd.read_excel(data, 'Sholl - Dendrites')
        dendrite_max_extent=df10.iloc[:,0].max()

        return dendrite_max_extent
    
    else:
        return np.nan
        
def make_param_df(datapath):

    data = e2p(datapath)
    sheetnames = get_sheet_names(data)

    Parameter_dict ={
#                    'Recording_Date': get_recording_date(datapath),
                    'Axon_node_total': Axon_node_total(data,sheetnames=sheetnames), 
                    'Axon_length': Total_axon_length(data,sheetnames=sheetnames),
                    'Axon_SA': Total_axon_SA(data,sheetnames=sheetnames),
                    'Axon_SA': Total_axon_SA(data,sheetnames=sheetnames),
                    'Axon_length_to_SA_ratio': Axon_len_SA_ratio(data,sheetnames=sheetnames),
                    'Highest_order_axon_segment': Highest_order_axon_segment(data,sheetnames=sheetnames),
                    'Axon_polar_angle': Axon_av_polar_angle(data,sheetnames=sheetnames), 
                    'Axon_polar_angle_std': Axon_std_polar_angle(data,sheetnames=sheetnames),
                    'Axon_local_angle': Axon_av_local_angle(data,sheetnames=sheetnames),
                    'Axon_local_angle_std': Axon_std_local_angle(data,sheetnames=sheetnames),
                    'Axon_spline_angle': Axon_av_spline_angle(data,sheetnames=sheetnames),
                    'Axon_spline_angle_std': Axon_std_spline_angle(data,sheetnames=sheetnames),
                    'Axon_segment_tortuosity': Axon_segment_av_tortuosity(data,sheetnames=sheetnames),
                    'Axon_segment_tortuosity_std': Axon_segment_std_tortuosity(data,sheetnames=sheetnames),
                    'Axon_segment_length': Axon_segment_av_length(data,sheetnames=sheetnames),
                    'Axon_segment_length_std': Axon_segment_std_length(data,sheetnames=sheetnames), 
                    'Axon_node_tortuosity': Axon_nodes_av_tortuosity(data,sheetnames=sheetnames),
                    'Axon_node_tortuosity_std': Axon_nodes_std_tortuosity(data,sheetnames=sheetnames),
                    'Axon_sholl_number': Axon_sholl_number(data,sheetnames=sheetnames),
                    'Axon_sholl_length_100': Axon_sholl_length_100(data,sheetnames=sheetnames),
                    'Axon_sholl_length_200': Axon_sholl_length_200(data,sheetnames=sheetnames),
                    'Axon_sholl_length_300': Axon_sholl_length_300(data,sheetnames=sheetnames),
                    'Axon_sholl_length_density': Axon_sholl_length_density(data,sheetnames=sheetnames),
                    'Axon_sholl_node_density': Axon_sholl_node_density(data,sheetnames=sheetnames),
                    'Dendrite_number': Dendrite_number(data,sheetnames=sheetnames),  
                    'Dendrite_node_total': Dendrite_node_total(data,sheetnames=sheetnames),  
                    'Total_dendrite_length': Total_dendrite_length(data,sheetnames=sheetnames), 
                    'Average_dendrite_length': Av_dendrite_length(data,sheetnames=sheetnames), 
                    'Dendrite_SA': Total_dendrite_SA(data,sheetnames=sheetnames), 
                    'Dendrite_length_to_SA_ratio': Dendrite_len_SA_ratio(data,sheetnames=sheetnames),  
                    'Highest_order_dendrite_segment': Highest_order_dendrite_segment(data,sheetnames=sheetnames), 
                    'Dendrite_polar_angle': Dendrite_av_polar_angle(data,sheetnames=sheetnames),  
                    'Dendrite_polar_angle_std': Dendrite_std_polar_angle(data,sheetnames=sheetnames), 
                    'Dendrite_local_angle': Dendrite_av_local_angle(data,sheetnames=sheetnames), 
                    'Dendrite_local_angle_std': Dendrite_std_local_angle(data,sheetnames=sheetnames), 
                    'Dendrite_spline_angle': Dendrite_av_spline_angle(data,sheetnames=sheetnames), 
                    'Dendrite_spline_angle_std': Dendrite_std_spline_angle(data,sheetnames=sheetnames), 
                    'Dendrite_segment_tortuosity': Dendrite_segment_av_tortuosity(data,sheetnames=sheetnames), 
                    'Dendrite_segment_tortuosity_std': Dendrite_segment_std_tortuosity(data,sheetnames=sheetnames), 
                    'Dendrite_segment_length': Dendrite_segment_av_length(data,sheetnames=sheetnames), 
                    'Dendrite_segment_length_std': Dendrite_segment_std_length(data,sheetnames=sheetnames),  
                    'Dendrite_node_tortuosity': Dendrite_nodes_av_tortuosity(data,sheetnames=sheetnames), 
                    'Dendrite_node_tortuosity_std': Dendrite_nodes_std_tortuosity(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_number': Dendrite_sholl_number(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_length_50': Dendrite_sholl_length_50(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_length_100': Dendrite_sholl_length_100(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_length_150': Dendrite_sholl_length_150(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_length_density': Dendrite_sholl_length_density(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_node_density': Dendrite_sholl_node_density(data,sheetnames=sheetnames), 
                    'Dendrite_sholl_number_10': Dendrite_sholl_number_10(data, sheetnames=sheetnames),
                    'Dendrite_sholl_number_50': Dendrite_sholl_number_50(data, sheetnames=sheetnames),
                    'Dendrite_sholl_number_100': Dendrite_sholl_number_100(data, sheetnames=sheetnames),
                    'Dendrite_sholl_number_200': Dendrite_sholl_number_200(data, sheetnames=sheetnames),
                    'Dendrite_sholl_number_300': Dendrite_sholl_number_300(data, sheetnames=sheetnames),
                    'Soma_perimeter': Soma_perimeter(data,sheetnames=sheetnames),
                    'Soma_surace_area': Soma_SA(data,sheetnames=sheetnames), 
                    'Soma_volume': Soma_volume(data,sheetnames=sheetnames), 
                    'Soma_aspect_ration': Soma_aspect_ratio(data,sheetnames=sheetnames),  
                    'Soma_compactness_factor': Soma_compactness_factor(data,sheetnames=sheetnames), 
                    'Soma_form_factor': Somatic_form_factor(data,sheetnames=sheetnames), 
                    'Somatic_roundness': Somatic_roundness(data,sheetnames=sheetnames),
                    'Dendrite_max_extent': dendrite_max_extent(data, sheetnames=sheetnames)
}

    df_params = pd.DataFrame.from_dict(Parameter_dict, orient = 'index')
    df_params.columns = [get_mouseID(datapath)]
    return (df_params)